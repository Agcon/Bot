# -*- coding: utf-8 -*-
import re
from bs4 import BeautifulSoup
import requests
import openpyxl
from datetime import datetime, timedelta ГОООООООООЙДА!!!!
import vk_api
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api import VkUpload
import PIL.Image as Image
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
from googletrans import Translator

def normalize_date(a):
    if a < 10: a = "0" + str(a)
    a = str(a)
    return a

def shedule_parse():
    url = "https://www.mirea.ru/schedule/"
    page = requests.get(url)
    links = []
    allTags = []
    soup = BeautifulSoup(page.text, "html.parser")
    allTags = soup.findAll('a', class_= "uk-link-toggle")
    for i in range(len(allTags)):
        allTags[i] = allTags[i].get('href')
    for link in allTags:
        if 'ИИТ' in link and 'курс_21-22' in link:
            links.append(link)
    current_link = links[3]
    for x in links:
        with open("D:/PythonProjects/Oznakom/file.xlsx", 'wb') as f:
            resp = requests.get(current_link)
            f.write(resp.content)
            
def get_schedule(number_of_group, weekday, evenness, book):
    sheet = book
    num_cols = sheet.max_column  
    row_index = 2
    column_index = 6
    cell = sheet.cell(row = 2, column = 6).value
    subjects = []
    for column_index in range(6, num_cols):
        cell = sheet.cell(row = row_index, column = column_index).value
        if (cell == number_of_group):
            for new_row in range(row_index + 2 + weekday * 12 + evenness, row_index + 14 + weekday * 12, 2):
                subjects.append(sheet.cell(row = new_row, column = column_index).value)
    return subjects

def get_evenness(day):
    first_day = datetime(2022, 2, 7)
    current_week = day - first_day
    current_week = 1 + int(current_week.days) // 7
    current_week %= 2
    if (current_week == 1): return 0
    return 1

def get_day_of_week(day):
    DAYS = {"понедельник": 0, "вторник": 1, "среда": 2, "четверг": 3,"пятница": 4, "суббота": 5, "воскресенье": 6}
    print(day)
    if (day.lower() == "сегодня"): return datetime.today().weekday()
    if (day.lower() == "завтра"): return (datetime.today() + timedelta(1)).weekday()
    if (day.lower() == "эта неделя" or day.lower() == "следующая неделя"): return 0
    return DAYS[day.lower()]

def get_week_schedule(number_of_group, day, book):
    DAYS = {0: "Понедельник", 1: "Вторник", 2: "Среда", 3: "Четверг", 4: "Пятница", 5: "Суббота"}
    even = get_evenness(day)
    s = ""
    for i in range(6):
        s += DAYS[i] + ":" + "\n"
        schedule = get_schedule(number_of_group, i, even, book)
        for j in range(len(schedule)):
            s += str(j + 1) + ") "
            if (schedule[j] == None): s += "-----" + "\n"
            else: s += str(schedule[j]) + "\n"
    return s

def get_correct_schedule(day, number_of_group, book):
    day1 = get_day_of_week(day)
    if (day.lower() == "сегодня"):
        schedule = get_schedule(number_of_group, day1, get_evenness(datetime.today()), book)
        s = ""
        for i in range(len(schedule)):
            s += str(i + 1) + ") "
            if (schedule[i] == None): s += "--" + "\n"
            else: s += str(schedule[i]) + "\n"
        return s
    if (day.lower() == "завтра"):
        schedule = get_schedule(number_of_group, day1, get_evenness(datetime.today() + timedelta(1)), book)
        s = ""
        for i in range(len(schedule)):
            s += str(i + 1) + ") "
            if (schedule[i] == None): s += "--" + "\n"
            else: s += str(schedule[i]) + "\n"
        return s
    if (day.lower() == "эта неделя"):
        return get_week_schedule(number_of_group, datetime.today(), book)
    if (day.lower() == "следующая неделя"):
        return get_week_schedule(number_of_group, datetime.today() + timedelta(7), book)
    schedule = get_schedule(number_of_group, get_day_of_week(day), 0, book)
    s = "Расписание для нечётной недели:" + "\n"
    for i in range(len(schedule)):
            s += str(i + 1) + ") "
            if (schedule[i] == None): s += "--" + "\n"
            else: s += str(schedule[i]) + "\n"
    s += "Расписание для чётной недели:" + "\n"
    schedule = get_schedule(number_of_group, get_day_of_week(day), 1, book)
    for i in range(len(schedule)):
            s += str(i + 1) + ") "
            if (schedule[i] == None): s += "--" + "\n"
            else: s += str(schedule[i]) + "\n"
    return s

def call_keyboard_teacher(keyboard, vk, event):
    keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
    keyboard.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)
    ms = 'Выберите варианты'
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), keyboard=keyboard.get_keyboard(), message=ms)

def get_teacher_schedule(name, book):
    full_name = ''
    schedule = [[['--', '--', '--', '--', '--', '--'] for _ in range(6)] for i in range(2)]
    for i in range(4, book.max_row):
        for j in range(8, book.max_column):
            try:
                if name.lower() in book.cell(i, j).value.lower():
                    full_name = book.cell(i, j).value
                    schedule[(i+1)%2][(i - 4) // 12][(i - 2) // 2 % 6 - 1] = ", ".join([(book.cell(i, j-2).value), (book.cell(i, j-1).value), (book.cell(2, j-2).value), (book.cell(i, j+1).value)]) 
            except AttributeError:
                pass
    return [full_name, schedule]

def get_correct_teacher_schedule(name, book, day):
    full_name, schedule = get_teacher_schedule(name, book)
    information = ""
    evenness = get_evenness(day)
    weekday = day.weekday()
    for i in range(6):
        information += "{}) ".format(str(i + 1)) + schedule[evenness][weekday][i] + "\n"
    return [information, full_name]

def bofort_scale(speed):
    if 0 <= speed <= 0.2: return "штиль"
    elif speed <= 1.5: return "тихий"
    elif speed <= 3.3: return "легкий"
    elif speed <= 5.4: return "слабый"
    elif speed <= 7.9: return "умеренный"
    elif speed <= 10.7: return "свежий"
    elif speed <= 13.8: return "сильный"
    elif speed <= 17.1: return "крепкий"
    elif speed <= 20.7: return "очень крепкий"
    elif speed <= 24.4: return "шторм"

def rumb(deg):
    if deg <= 22.5 or deg >= 337.5: return "северный"
    elif deg <= 67.5: return "северо-восточный"
    elif deg <= 112.5: return "восточный"
    elif deg <= 157.5: return "юго-восточный"
    elif deg <= 202.5: return "южный"
    elif deg <= 247.5: return "юго-западный"
    elif deg <= 292.5: return "западный"
    return "северо-западный"

def call_keyboard_first(keyboard, vk, event):
    keyboard.add_button('Получить расписание', color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button('Получить погоду', color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button('Получить статистику по коронавирусу', color=VkKeyboardColor.POSITIVE)
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), keyboard=keyboard.get_keyboard(), message='Что вы хотите узнать?')

def first_message(vk, event):
    print('New from {}, text = {}'.format(event.user_id, event.text))
    vk.messages.send(
    user_id = event.user_id,
    random_id = get_random_id(),
    message = 'Привет, ' + \
    vk.users.get(user_id = event.user_id)[0]['first_name']
    )
    vk.messages.send(
    user_id = event.user_id,
    random_id = get_random_id(),
    message = 'Напиши БОТ, если хочешь посмотреть функции'
    )

def call_keyboard_schedule(keyboard, vk, event):
    keyboard.add_button('сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard.add_button('завтра', color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button('эта неделя', color=VkKeyboardColor.PRIMARY)
    keyboard.add_button('следующая неделя', color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button('какая неделя?', color=VkKeyboardColor.SECONDARY)
    keyboard.add_button('какая группа?', color=VkKeyboardColor.SECONDARY)
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), keyboard=keyboard.get_keyboard(), message='Выберите варианты')

def call_keyboard_weather(keyboard, vk, event):
    keyboard.add_button('сейчас', color=VkKeyboardColor.POSITIVE)
    keyboard.add_button('сегодня', color=VkKeyboardColor.PRIMARY)
    keyboard.add_button('завтра', color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button('на 5 дней', color=VkKeyboardColor.NEGATIVE)
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), keyboard=keyboard.get_keyboard(), message='Выберете варианты')

def get_weather_now():
    translator = Translator()
    weather_key = "1e7d1c94703c5b863a60ea656e79de92"
    response = requests.get("http://api.openweathermap.org/data/2.5/weather?q=moscow&appid=1e7d1c94703c5b863a60ea656e79de92&lang=ru&units=metric")
    info = response.json()
    information = ""
    result = translator.translate(info["weather"][0]["main"], src="en", dest="ru").text
    information += "Погода в Москве: " + result + "\n"
    result = info["weather"][0]["description"]
    information += result + ", " + "температура: " + str(round(info["main"]["temp_min"])) + " - " + str(round(info["main"]["temp_max"])) + " °C" + "\n"
    information += "Давление: " + str(info["main"]["pressure"]) + " мм.рт.ст., " + "влажность: " + str(info["main"]["humidity"]) + "%" + "\n"
    speed = info["wind"]["speed"]
    deg = info["wind"]["deg"]
    information += "Ветер: " + bofort_scale(speed) + ", " + str(speed) + " м/с, " + rumb(deg)
    return information

def weather_in_time(response):
    image = requests.get("http://openweathermap.org/img/wn/{}@2x.png".format(response['weather'][0]['icon']), stream=True)

    information = ""
    result = response["weather"][0]["description"]
    information += result + ", " + "температура: " + str(round(response["main"]["temp_min"])) + " - " + str(round(response["main"]["temp_max"])) + " °C" + "\n"
    information += "Давление: " + str(response["main"]["pressure"]) + " мм.рт.ст., " + "влажность: " + str(response["main"]["humidity"]) + "%" + "\n"
    speed = response["wind"]["speed"]
    deg = response["wind"]["deg"]
    information += "Ветер: " + bofort_scale(speed) + ", " + str(speed) + " м/с, " + rumb(deg)

    return [image, information]

def get_weather_today(day, vk_session, vk, event):
    response = requests.get("http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid=1e7d1c94703c5b863a60ea656e79de92&lang=ru&units=metric")
    info = response.json()
    ms = "Погода в Москве " + day
    data = datetime.today().date()
    information = ""
    if (day == "завтра"): data = (datetime.today() + timedelta(1)).date()
    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
    counter = 0
    for i in range(len(info["list"])):
        if str(data) in info["list"][i]["dt_txt"]:
            if "6:00:00" in info["list"][i]["dt_txt"]:
                response = info["list"][i]
                collect = weather_in_time(response)
                information += "УТРО:\n"
                attachments = collect[0]
                information += collect[1] + "\n"
                counter += 1
                with open("file1.png", "wb") as f:
                    f.write(attachments.content)
            if "12:00:00" in info["list"][i]["dt_txt"]:
                response = info["list"][i]
                collect = weather_in_time(response)
                attachments = collect[0]
                information += "ДЕНЬ:\n"
                information += collect[1] + "\n"
                counter += 1
                with open("file2.png", "wb") as f:
                    f.write(attachments.content)
            if "18:00:00" in info["list"][i]["dt_txt"]:
                response = info["list"][i]
                collect = weather_in_time(response)
                attachments = collect[0]
                information += "ВЕЧЕР:\n"
                information += collect[1] + "\n"
                counter += 1
                with open("file3.png", "wb") as f:
                    f.write(attachments.content)
            if "21:00:00" in info["list"][i]["dt_txt"]:
                response = info["list"][i]
                collect = weather_in_time(response)
                attachments = collect[0]
                information += "НОЧЬ:\n"
                information += collect[1] + "\n"
                counter += 1
                with open("file4.png", "wb") as f:
                    f.write(attachments.content)
                if counter == 4:
                    img = Image.new('RGBA', (400, 100))
                    img1 = Image.open("file1.png")
                    img2 = Image.open("file2.png")
                    img3 = Image.open("file3.png")
                    img4 = Image.open("file4.png")
                    img.paste(img1, (0, 0))
                    img.paste(img2, (100, 0))
                    img.paste(img3, (200, 0))
                    img.paste(img4, (300, 0))
                elif counter == 3:
                    img = Image.new('RGBA', (300, 100))
                    img2 = Image.open("file2.png")
                    img3 = Image.open("file3.png")
                    img4 = Image.open("file4.png")
                    img.paste(img2, (0, 0))
                    img.paste(img3, (100, 0))
                    img.paste(img4, (200, 0))
                elif counter == 2:
                    img = Image.new('RGBA', (200, 100))
                    img3 = Image.open("file3.png")
                    img4 = Image.open("file4.png")
                    img.paste(img3, (0, 0))
                    img.paste(img4, (100, 0))
                elif counter == 1:
                    img = Image.new('RGBA', (100, 100))
                    img4 = Image.open("file4.png")
                    img.paste(img4, (0, 0))
                img = img.save("image.png")
                upload = VkUpload(vk_session)
                photo = upload.photo_messages(photos="image.png")[0]
                attachments = ("photo{}_{}".format(photo["owner_id"], photo['id']))
                vk.messages.send(
                    user_id=event.user_id,
                    random_id=get_random_id(),
                    attachment=attachments,
                    message="\n")
                vk.messages.send(user_id=event.user_id, random_id=get_random_id(), message=information)
                break

def get_weather_in_5_days(vk_session, vk, event):
    response = requests.get("http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid=1e7d1c94703c5b863a60ea656e79de92&lang=ru&units=metric")
    info = response.json()
    date1 = datetime.today() + timedelta(1)
    date1 = normalize_date(date1.day) + "." + normalize_date(date1.month)
    date2 = datetime.today() + timedelta(5)
    date2 = normalize_date(date2.day) + "." + normalize_date(date2.month)
    ms = "Погода в Москве c " + date1 + " по " + date2
    vk.messages.send(user_id=event.user_id, random_id=get_random_id(), message=ms)
    day = []
    night = []
    images = []
    counter = 0
    for i in info["list"]:
        if (counter == 5): break
        if '03:00:00' in i["dt_txt"]:
            night.append(str(round(i['main']['temp'])) + ' °C')
        elif '15:00:00' in i["dt_txt"]:
            counter += 1
            day.append(str(round(i['main']['temp'])) + ' °C')
            images.append(i['weather'][0]['icon'])
    counter = 0
    for i in images:
        counter += 1
        with open("file{}.png".format(counter), "wb") as f:
            image=requests.get("http://openweathermap.org/img/wn/{}@2x.png".format(i), stream=True)
            f.write(image.content)
    img = Image.new('RGBA', (500, 100))
    img1 = Image.open("file1.png")
    img2 = Image.open("file2.png")
    img3 = Image.open("file3.png")
    img4 = Image.open("file4.png")
    img5 = Image.open("file4.png")
    img.paste(img1, (0, 0))
    img.paste(img2, (100, 0))
    img.paste(img3, (200, 0))
    img.paste(img4, (300, 0))
    img.paste(img5, (400, 0))
    img = img.save("image.png")
    upload = VkUpload(vk_session)
    photo = upload.photo_messages(photos="image.png")[0]
    attachments = ("photo{}_{}".format(photo["owner_id"], photo['id']))
    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        attachment=attachments,
        message="\n")
    ms = "\n" + 'ДЕНЬ: '
    for i in day:
        ms += i + " // "
    ms += "\n" + "НОЧЬ: "
    for i in night:
        ms += i + " // "
    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        message=ms)

def get_coronavirus_stat(vk, vk_session, event):
    response = requests.get("https://coronavirusstat.ru/country/russia/")
    soup = BeautifulSoup(response.text, "html.parser")
    result = soup.findAll('table')[0].find("tbody").findAll("td")
    s = soup.findAll('body')[0].find("h6").find('strong').text + '\n'
    k = 0
    for i in result[0]:
        if k == 0:
            s += "Активных: " + str(i)
        elif k == 1:
            s += "({} за сегодня)".format(i.text) + '\n'
        else:
            break
        k += 1
    k = 0
    for i in result[1]:
        if k == 0:
            s += "Вылечено: " + str(i)

        elif k == 1:
            s += "({} за сегодня)\n".format(i.text)
        else:
            break
        k += 1
    k = 0
    for i in result[2]:
        if k == 0:
            s += "Умерло: " + str(i)
        elif k == 1:
            s += "({} за сегодня)\n".format(i.text)
        else:
            break
        k += 1
    k = 0
    for i in result[3]:
        if k == 0:
            s += "Случаев: " + str(i)

        elif k == 1:
            s += "({} за сегодня)".format(i.text)
        else:
            break
        k += 1
    result = soup.findAll('table')[0].find("tbody").findAll("td", {"class": "d-none d-sm-block"})
    infected = []
    k = 0
    for i in result:
        if k < 10:
            if i.find("span", {"class": "badge badge-danger"}):
                infected.append(int(i.find("span", {"class": "badge badge-danger"}).text))
                k += 1
        else:
            break
    
    result = soup.findAll('table')[0].find("tbody").findAll("span", {"class": "badge badge-success"})

    cured = []
    k = 0
    for i in result:
        if k < 20 and k % 2 == 1:
            print(i.text)
            cured.append(int(i.text))
        elif k > 20:
            break
        k += 1
    result = soup.findAll('table')[0].find("tbody").findAll("th")
    data = []
    k = 0
    for i in result:
        if k < 10:
            print(i.text)
            data.append(i.text[:5])
        else:
            break
        k += 1
    a = np.array([cured, infected])
    barWidth = 0.25
    fig = plt.subplots()

    br1 = np.arange(len(cured))
    br2 = [x + barWidth for x in br1]

    plt.bar(br1, cured, color='b', width=barWidth,
            edgecolor='grey', label='Выздоровевшие')
    plt.bar(br2, infected, color='g', width=barWidth,
            edgecolor='grey', label='Заболевшие')

    plt.xlabel('', fontweight='bold', fontsize=15)
    plt.ylabel('Кол-во', fontweight='bold', fontsize=15)
    plt.xticks([r + barWidth for r in range(len(cured))],
                data)

    plt.legend()
    plt.savefig("grafic.png")

    upload = VkUpload(vk_session)
    photo = upload.photo_messages(photos="grafic.png")[0]
    attachments = []
    attachments.append("photo{}_{}".format(photo["owner_id"], photo['id']))
    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        attachment=attachments[0],
        message=s)

def get_coronavirus_stat_by_region(region, vk, event):
    response = requests.get("https://coronavirusstat.ru")
    soup = BeautifulSoup(response.text, "html.parser")
    result = soup.findAll('div', {'class': "row border border-bottom-0 c_search_row"})
    a = ''
    region = region.capitalize()
    for i in range(len(result)):
        if region in result[i].find('a').text:
            s = result[i].findAll("span", {"class": "dline"})
            a += "Активных: "+s[0].text+"\n"
            a += "Вылечено: "+s[1].text+"\n"
            a += "Умерло: "+s[2].text+"\n"
            s = result[i].findAll("div", {"class": "h6 m-0"})
            a += "Заразилось: "+s[0].text[2:]
            break
    if a == "":
        vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        message="Регион не найден")
    else:
        vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        message=a[:len(a) - 11])

def main():
    #shedule_parse()
    book1 = openpyxl.load_workbook("D:/PythonProjects/Oznakom/file1.xlsx")
    book2 = openpyxl.load_workbook("D:/PythonProjects/Oznakom/file2.xlsx")
    book3 = openpyxl.load_workbook("D:/PythonProjects/Oznakom/file3.xlsx")
    book1 = book1.active
    book2 = book2.active
    book3 = book3.active
    token = "8795f2314abfac327507b600196e7b594645de3c7f744b5acd5058e1f36b56b84fbed61db31b092b052fe"
    vk_session = vk_api.VkApi(token=token)
    vk = vk_session.get_api()
    longpoll = VkLongPoll(vk_session)
    last_message = ""
    current_group = "ИКБО-08-21"
    #schedule_variants = ["сегодня", "завтра", "эта неделя", "следующая неделя"]
    flag_schedule = False
    flag_weather = False
    flag_teacher = False

    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW:
            full_message = list(event.text.split())

        if event.type == VkEventType.MESSAGE_NEW and (event.text.lower() == "start" or event.text.lower() == "начать"):
            ms = 'Для начала работы с ботом напишите привет. Функции бота можно посомотреть, введя БОТ' + "\n" + 'Узнать расписание для конкретной группы можно, введя запрос в формате БОТ + "номер группы" или же БОТ "день недели" "номер группы" ' + "\n" + 'При вводе Коронавирус "название региона" выведется статистика по коронавирусу в текущем регионе'
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
        
        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "привет" and event.to_me:
            first_message(vk, event)

        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "бот":
            keyboard = VkKeyboard(one_time=True)
            call_keyboard_first(keyboard, vk, event)

        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "получить расписание":
            keyboard = VkKeyboard(one_time=True)
            call_keyboard_schedule(keyboard, vk, event)
            flag_schedule = True

        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "получить погоду":
            keyboard = VkKeyboard(one_time=True)
            call_keyboard_weather(keyboard, vk, event)
            flag_weather = True

        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 1 and full_message[0][0].lower() == "и":
            current_group = full_message[0].upper()
            ms = "Я запомнил, что ты из группы " + full_message[0].upper()
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)

        elif event.type == VkEventType.MESSAGE_NEW and flag_schedule and event.to_me:
            if event.text != "какая неделя?" and event.text != 'какая группа?':
                if current_group[-1] == '1':
                    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=get_correct_schedule(event.text, current_group, book1))
                elif current_group[-1] == '0':
                    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=get_correct_schedule(event.text, current_group, book2))
                elif current_group[-1] == '9':
                    vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=get_correct_schedule(event.text, current_group, book3))
            elif event.text == "какая неделя?":
                first_day = datetime(2022, 2, 7)
                current_week = datetime.today() - first_day
                current_week = 1 + int(current_week.days) // 7
                ms = "Сейчас идёт " + str(current_week) + " неделя"
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            else:
                ms = "Показываю расписание группы " + current_group
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            flag_schedule = False

        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 2 and full_message[0].lower() == "бот":
            if full_message[1][0].lower() == "и":
                current_group = full_message[1].upper()
                ms = "Выбрана группа " + current_group
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
                ms = "Получить расписание"
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            else:
                if current_group[-1] == '1':
                    ms = get_correct_schedule(full_message[1].lower(), current_group, book1)
                elif current_group[-1] == '0':
                    ms = get_correct_schedule(full_message[1].lower(), current_group, book2)
                elif current_group[-1] == '9':
                    ms = get_correct_schedule(full_message[1].lower(), current_group, book3)
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)

        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 3 and full_message[0].lower() == "бот":
            if full_message[2] == '1':
                ms = get_correct_schedule(full_message[1].lower(), full_message[2].upper(), book1)
            if full_message[2] == '0':
                ms = get_correct_schedule(full_message[1].lower(), full_message[2].upper(), book2)
            if full_message[2] == '9':
                ms = get_correct_schedule(full_message[1].lower(), full_message[2].upper(), book3)
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
        
        elif event.type == VkEventType.MESSAGE_NEW  and event.text.lower() == "в чём смысл жизни?":
            ms = "Эта жизнь хуйня такая, нихуя ты не поймёшь, каждый день охуеваешь, охуевшим и помрёшь!"
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
        
        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 2 and full_message[0].lower() == "найти":
            keyboard = VkKeyboard(one_time=True)
            call_keyboard_teacher(keyboard, vk, event)
            name_of_teacher = full_message[1]
            flag_teacher = True

        elif event.type == VkEventType.MESSAGE_NEW and flag_teacher and event.to_me:
            DAYS = {0: "Понедельник", 1: "Вторник", 2: "Среда", 3: "Четверг", 4: "Пятница", 5: "Суббота"}
            if event.text.lower() == "на сегодня":
                ms, namee = get_correct_teacher_schedule(name_of_teacher, book1, datetime.today())
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message="Показываю расписание на преподавателя " + namee)
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            elif event.text.lower() == "на завтра":
                ms, namee = get_correct_teacher_schedule(name_of_teacher, book1, datetime.today() + timedelta(1))
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message="Показываю расписание на преподавателя " + namee)
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            elif event.text.lower() == "на эту неделю":
                date = datetime.today() - timedelta(datetime.today().weekday())
                ms, namee = get_correct_teacher_schedule(name_of_teacher, book1, datetime.today())
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message="Показываю расписание на преподавателя " + namee)
                ms = ''
                for i in range(6):
                    ms += DAYS[i] + ":" + "\n"
                    mmss, namee = get_correct_teacher_schedule(name_of_teacher, book1, date + timedelta(i))
                    ms += mmss
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            elif event.text.lower() == "на следующую неделю":
                date = datetime.today() - timedelta(datetime.today().weekday()) + timedelta(7)
                ms, namee = get_correct_teacher_schedule(name_of_teacher, book1, datetime.today())
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message="Показываю расписание на преподавателя " + namee)
                ms = ''
                for i in range(6):
                    ms += DAYS[i] + ":" + "\n"
                    mmss, namee = get_correct_teacher_schedule(name_of_teacher, book1, date + timedelta(i))
                    ms += mmss
                vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            flag_teacher = False

        elif event.type == VkEventType.MESSAGE_NEW and flag_weather and event.text.lower() == "сейчас":
            ms = get_weather_now()
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)
            flag_weather = False

        elif event.type == VkEventType.MESSAGE_NEW and flag_weather and (event.text.lower() == "сегодня" or event.text.lower() == "завтра"):
            get_weather_today(event.text.lower(), vk_session, vk, event)
            flag_weather = False
        
        elif event.type == VkEventType.MESSAGE_NEW and flag_weather and event.text.lower() == "на 5 дней":
            get_weather_in_5_days(vk_session, vk, event)
            flag_weather = False

        elif event.type == VkEventType.MESSAGE_NEW and event.text.lower() == "получить статистику по коронавирусу":
            get_coronavirus_stat(vk, vk_session, event)

        elif event.type == VkEventType.MESSAGE_NEW and len(full_message) == 2 and full_message[0].lower() == "коронавирус":
            get_coronavirus_stat_by_region(full_message[1].lower(), vk, event)

        elif event.type == VkEventType.MESSAGE_NEW and event.to_me:
            ms = "Неизвестная команда. Для списка команд напишите БОТ"
            vk.messages.send(user_id = event.user_id, random_id = get_random_id(), message=ms)

if __name__ == "__main__":
    main()
